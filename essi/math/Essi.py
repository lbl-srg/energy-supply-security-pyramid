#!/usr/bin/env python
# -*- coding: utf-8 -*-
#######################################################
# Class that implements the energy security metrics
#
#######################################################
import openpyxl
import numpy as np
import pandas as pd

class Essi(object):
    """ Class that implements the energy security metrics.

        Note that I, P, D, E, L, Sb and Sd must all have the same
        number of energy carriers, and these must be in the same order.
        For example, if the first column of I is electricity, then the first
        column of P must also be electricity, as these energy carriers
        are added, subtracted or compared for each carrier type.

    """

    def __init__(self, filename, sheet,
                 sigma="B16",
                 time="A20:A8779",
                 I="AY20:BM8779",
                 P="AI20:AW8779",
                 D="C20:Q8779",
                 E="BO20:CC8779",
                 L="S20:AG8779",
                 Sb="CE20:CS8779",
                 Sd="CU20:DI8779",
                 Sto_start="CE17:CS17",
                 Sto_end="CU18:DI18",
                 a="AY14:BM14",
                 c="C13:Q13",
                 n="C15:Q15"):

        # --------------------------
        # Class variables
        # self._libHome=os.path.abspath(".")
        self._filename = filename
        self._sheet = sheet

        # data frames
        time = self._get_pandas_frame(time)
        time_list = time[time.columns.to_list()[0]]
        self._time_step = time_list[1] - time_list[0]
        self._sigma = self._get_cell_value(sigma)
        self._I = self._get_pandas_frames(I)
        self._P = self._get_pandas_frames(P)
        self._D = self._get_pandas_frames(D)
        self._E = self._get_pandas_frames(E)
        self._L = self._get_pandas_frames(L)
        self._Sb = self._get_pandas_frames(Sb).fillna(0)
        self._Sd = self._get_pandas_frames(Sd).fillna(0)
        self._Sto_start = (self._get_pandas_frames(Sto_start).fillna(0)).iloc[0].to_list()
        self._Sto_end = (self._get_pandas_frames(Sto_end).fillna(0)).iloc[0].to_list()
        self._a = (self._get_pandas_frames(a)).iloc[0].to_list()
        self._c = (self._get_pandas_frames(c)).iloc[0].to_list()
        self._n = (self._get_pandas_frames(n)).iloc[0].to_list()


        if self._I.shape != self._P.shape:
            raise RuntimeError("I and P must have the same dimensions.")
        if self._I.shape != self._D.shape:
            raise RuntimeError("I and D must have the same dimensions.")
        if self._I.shape != self._E.shape:
            raise RuntimeError("I and E must have the same dimensions.")
        if self._I.shape != self._L.shape:
            raise RuntimeError("I and L must have the same dimensions.")
        if self._I.shape != self._Sb.shape:
            raise RuntimeError("I and Sb must have the same dimensions.")
        if self._I.shape != self._Sd.shape:
            raise RuntimeError("I and Sd must have the same dimensions.")
        if self._I.shape[1] != len(self._Sto_start):
            raise RuntimeError("I and Sto_start must have the same number of energy flows.")
        if self._I.shape[1] != len(self._Sto_end):
            raise RuntimeError("I and Sto_end must have the same number of energy flows.")
        if self._I.shape[1] != len(self._c):
            raise RuntimeError("I and c must have the same number of energy flows.")
        if self._I.shape[1] != len(self._a):
            raise RuntimeError("I and a must have the same number of energy flows.")

        if not np.isscalar(self._sigma):
            raise RuntimeError("sigma must be a single cell.")

    def _get_cell_value(self, cell_id):
        """
        Get the value of an individual cell

        :param: The id of the cell, such as "A1" for the first cell of the sheet
        """
        from openpyxl import load_workbook
        wb = load_workbook(self._filename, read_only=True, data_only=True)
        ws = wb[self._sheet]
        return ws[cell_id].value

    def _get_pandas_frames(self, range_string):
        """ Low level function to return Pandas data frame.

        :param: range_string: Range of values if xls file, such as "A2:A4", or "A2:A4,"C2:C4"".
        """
        import pandas as pd
        frames = []
        for rs in range_string.split(","):
            fr=self._get_pandas_frame(rs)
            frames.append(fr)
        frame = pd.concat(frames, axis=1)
        return frame

    def _get_pandas_frame(self, range_string):
        """ Low level function to return Pandas data frame.
        """
        from openpyxl import load_workbook
        import pandas

        def load_workbook_range(range_string, ws):
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_interval
            import re

            col_start, col_end = re.findall("[A-Z]+", range_string)

            data_rows = []
            for row in ws[range_string]:
                data_rows.append([cell.value for cell in row])
            return pandas.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))

        wb = load_workbook(
                filename=self._filename,
                read_only=True,
                data_only=True)
        ws = wb[self._sheet]
        vals = load_workbook_range(range_string, ws)
        return vals

    @staticmethod
    def integrate_by_columns(data_frame, time_step):
        """
        Integrate each column individually and return an array of
        the integral of each column.

        Note that this function assumes each entry to have
        the same time step dt
        """
        sums = []
        for c in data_frame.columns.to_list():
            sums.append(time_step * sum(data_frame[c]))
        return sums

    @staticmethod
    def integrate_all_columns(data_frame, time_step):
        return sum(Essi.integrate_by_columns(data_frame, time_step))

    @staticmethod
    def get_d(P, time_step, setM, carrier_type):
        phis = Essi.get_phis(P, time_step, setM)
        n = len(phis)
        if n == 0:
            raise RuntimeError(f"Attempting to compute diversity index, but all energy carriers for {carrier_type} are zero.")
        summand = 0
        iC = 0
        for phi in phis:
            if phis[iC] < 1E-4:
                raise RuntimeError(f"Error when calculating d. phi[{iC}] = {phis[iC]}, but it need to be positive to compute the log.")
            iC += 1
            summand += phi * np.log(phi)

        d = - summand / np.log(n)
        return d


    @staticmethod
    def get_columns(P, setM):
        return np.transpose(np.array([P[:,i] for i in setM]))


    @staticmethod
    def get_phis(P, time_step, setM):
        all_integrals = Essi.integrate_by_columns(P, time_step)
        integrals = [all_integrals[i] for i in setM]
        denominator = sum(integrals)
        phi = list(np.multiply(1/denominator, integrals))
        return phi

    @staticmethod
    def get_setM(P):
        intCols = Essi.integrate_by_columns(P, 1)
        m = [x for x in range(len(intCols)) if intCols[x] > 0.0001]
        return m

    def get_SPG(self):
        """
        Returns the self-production grade SPG
        """

        num = []

        setM = Essi.get_setM(self._P)

        P = Essi.integrate_by_columns(self._P, self._time_step)
        L = Essi.integrate_by_columns(self._L, self._time_step)

        D = Essi.integrate_by_columns(self._D, self._time_step)

        num = 0
        for i in setM:
            num += (P[i]-L[i]+self._Sto_end[i]-self._Sto_start[i])

        den = sum(D)

        SPG = Essi.get_d(self._P, self._time_step, setM, "P") * num / den
        return SPG

    def get_SAG(self):
        """
        Returns the adequacy SAG
        """
        def get_f_i():
            P = self._P.to_numpy()
            I = self._I.to_numpy()
            E = self._E.to_numpy()
            Sd = self._Sd.to_numpy()
            Sb = self._Sb.to_numpy()
            L = self._L.to_numpy()
            D = self._D.to_numpy()

            lhs = P + I + E + Sd - Sb - L
            rhs = D

            retB = lhs >= rhs - np.full(rhs.shape, 1E-6)
            #retB = abs(lhs >= rhs)
            ret = retB.astype(float)

            return ret

        f = get_f_i()

        D = self._D.to_numpy()
        nSte = D.shape[0]
        nFlo = D.shape[1]
        sum_D_over_j = np.zeros(D.shape[0])
        for i in range(nFlo):
            sum_D_over_j += D[:,i]

        # Array with each element being an energy carrier i, and the array contains the integral
        a = np.zeros(nFlo)
        for i in range(nFlo):
            # Integrate over each time step and divide each term by sum_j D_j(t)
            for k in range(nSte): # k is the time step
                a[i] += f[k, i] * D[k, i] / sum_D_over_j[k]
            a[i] = a[i] / float(nSte)

        summand = self._c * a
        SAG = np.sum(summand)

        return SAG

    def get_AUG(self):
        """
        Get the autonomy grade AUG
        """

        # Build set of non-zero contributions
        setM_I = Essi.get_setM(self._I)
        setM_D = Essi.get_setM(self._D)
        # Intersection of the two sets
        setM = [x for x in setM_I if x in setM_D]

        I_inM = Essi.get_columns(self._I.to_numpy(), setM)

        # Compute phi (formerly called alpha)
        int_I = self._time_step * sum(sum(self._I.to_numpy()))
        int_D = self._time_step * sum(sum(self._D.to_numpy()))
        int_L = self._time_step * sum(sum(self._L.to_numpy()))
        phi = int_I/(int_D+int_L)

        # Compute theta
        D_inM = Essi.get_columns(self._D.to_numpy(), setM)
        L_inM = Essi.get_columns(self._L.to_numpy(), setM)
        # Add D and L element-wise
        D_L_inM = np.add(D_inM, L_inM)
        int_D_L_inM = self._time_step * sum(D_L_inM)


        # Integral \int_T I_i(t) dt
        # int_I_i is a vector whose elements are the integrals of each energy carrier i
        int_I_i = self._time_step * sum(I_inM)

        # Sum of all integrals, e.g., the denominator sum_j \int_T I_j(t) dt
        int_I = np.sum(int_I_i)

        # Compute a'
        cardSetM = np.size(I_inM, 1)
        theta = np.zeros(cardSetM)
        stoEnd_inM = [self._Sto_end[i] for i in setM]
        n_inM = [self._n[i] for i in setM]
        for i in range(cardSetM):
            if n_inM[i] < 1E-4:
                raise ValueError(f"Error in AUG calculation n[{i}] must not be zero. Received n[{i}]={n_inM[i]}.")

            if abs(int_D_L_inM[i]) < 1E-4:
                raise ValueError(f"Error in AUG calculation as imports of carrier {i} is zero, but this carrier is in M(D,T).")
            theta[i] = stoEnd_inM[i] / n_inM[i] / int_D_L_inM[i]
            if theta[i] > 1:
                theta[i] = 1

        # Compute a'
        a_inM = [self._a[i] for i in setM]
        aP_i = np.zeros(cardSetM)
        for i in range(cardSetM):
            aP_i[i] = (a_inM[i] + (1-a_inM[i]) * theta[i])*int_I_i[i]/int_I
        aP = sum(aP_i)

        # Compute AUG
        d = Essi.get_d(self._I, self._time_step, setM, "I")
        AUG = (1-phi) + aP * d * phi

        return AUG

    def get_SSG(self):
        """
        Returns the Self-Sufficiency Grad
        """
        def get_g_i():
            P = self._P.to_numpy()
            Sd = self._Sd.to_numpy()
            Sb = self._Sb.to_numpy()
            L = self._L.to_numpy()
            E = self._E.to_numpy()
            D = self._D.to_numpy()

            lhs = P + Sd - Sb - L - E
            rhs = D

            retB = lhs >= rhs - np.full(rhs.shape, 1E-6)
            #retB = abs(lhs >= rhs)
            ret = retB.astype(float)

            return ret

        g = get_g_i()

        D = self._D.to_numpy()
        nSte = D.shape[0]
        nFlo = D.shape[1]
        sum_D_over_j = np.zeros(D.shape[0])
        for i in range(nFlo):
            sum_D_over_j += D[:,i]
        # Compute integral, for each energy carrier i
        num = g * D
        # Array with each element being an energy carrier i, and the array contains the integral
        a = np.zeros(nFlo)
        for i in range(nFlo):
            # Integrate over each time step and divide each term by sum_j D_j(t)
            a1 = num[:,i] / sum_D_over_j
            a[i] = sum(a1) / nSte

        SSG = np.sum(a)

        return SSG

    def get_AUT(self):
        """
        Get the Autarky Grad
        """
        return self._sigma * self.get_SSG()

    def get_ESSI(self, w):
        """
        This functions returns the energy supply security index

        :param w[]: Array of weights w1, ..., w5 for [SPG, SAG, AUG, SSG, AUT]
        :return: double The energy supply security index

        Usage: Type
            >>> import math.Essi as c
            >>> m = c.Essi()
            >>> m.ESSI(w=[0.1, 0.2, 0.3, 0.35, 0.05])
        """
        return  (w[0] * self.get_SPG()
                 + w[1] * self.get_SAG()
                 + w[2] * self.get_AUG()
                 + w[3] * self.get_SSG()
                 + w[4] * self.get_AUT())/sum(w)
